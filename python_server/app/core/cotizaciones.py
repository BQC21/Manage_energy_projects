# backend/app/core/cotizaciones.py
from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any, Dict, Literal, Optional, Union

from openpyxl import load_workbook

from .configs import build_config
from .functions import construir_gastos, total_cot, total_cot_value, count_nro_panels
from .functions import cf_table, finantial_table, generar_graficas, _safe_div
from .pdf.quote import build_reporte_pdf
from .pdf.finantial import build_reporte_pdf_finantial

warnings.filterwarnings("ignore")

ReportType = Literal["quote", "finantial"]


def _resolve_root_path(cfg: dict) -> Path:
    root_cfg = cfg.get("data_roots", {}).get("ROOT_PATH")
    if not root_cfg:
        raise ValueError("CFG['data_roots']['ROOT_PATH'] está vacío.")
    return Path(root_cfg).resolve()


def _resolve_output_pdf_path(
    root_path: Path,
    output_pdf_path: Optional[Union[str, Path]],
) -> Path:
    # Conserva el comportamiento anterior cuando no llega path explícito.
    if output_pdf_path is None:
        output_pdf_path = root_path / "Outputs" / "reporte_final.pdf"
    else:
        output_pdf_path = Path(output_pdf_path)

    output_pdf_path = output_pdf_path.resolve()
    output_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    return output_pdf_path


def _build_quote_pdf(
    cfg: dict,
    ws,
    dict_datos: dict,
    assets: dict,
    root_path: Path,
    output_pdf_path: Path,
    dynamic_layout: Optional[Dict[str, Any]] = None,
) -> Path:
    df_gastos, df_equipos, df_principal = construir_gastos(
        ws,
        cfg["parametros_busqueda_BUDGET"]["minimo_vector_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["maximo_vector_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["minimo_vector_MO"],
        cfg["parametros_busqueda_BUDGET"]["maximo_vector_MO"],
        cfg["parametros_busqueda_BUDGET"]["column_equipos_desc"],
        cfg["parametros_busqueda_BUDGET"]["column_equipos_cant"],
        cfg["parametros_busqueda_BUDGET"]["column_equipos_uni"],
        cfg["parametros_busqueda_BUDGET"]["column_precio_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["row_precio_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["column_precio_MO"],
        cfg["parametros_busqueda_BUDGET"]["row_precio_MO"],
        dynamic_layout=dynamic_layout,
    )

    resumen = total_cot(df_principal)

    build_reporte_pdf(
        output_pdf_path=output_pdf_path,
        root_path=root_path,
        assets_logo=assets["logo"],
        assets_firma=assets["firma"],
        DF_cotizacion=df_principal,
        DF_budget=df_gastos,
        DF_equipos=df_equipos,
        resumen=resumen,
        dict_informe=cfg["datos_informe"],
        dict_datos=dict_datos,
    )

    return output_pdf_path


def compute_quote_metrics(
    cfg: dict,
    ws,
    dynamic_layout: Optional[Dict[str, Any]] = None,
) -> dict:
    _, df_equipos, df_principal = construir_gastos(
        ws,
        cfg["parametros_busqueda_BUDGET"]["minimo_vector_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["maximo_vector_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["minimo_vector_MO"],
        cfg["parametros_busqueda_BUDGET"]["maximo_vector_MO"],
        cfg["parametros_busqueda_BUDGET"]["column_equipos_desc"],
        cfg["parametros_busqueda_BUDGET"]["column_equipos_cant"],
        cfg["parametros_busqueda_BUDGET"]["column_equipos_uni"],
        cfg["parametros_busqueda_BUDGET"]["column_precio_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["row_precio_EQUIPOS"],
        cfg["parametros_busqueda_BUDGET"]["column_precio_MO"],
        cfg["parametros_busqueda_BUDGET"]["row_precio_MO"],
        dynamic_layout=dynamic_layout,
    )
    resumen = total_cot(df_principal)
    return {
        "price": total_cot_value(resumen),
        "nro_panels": count_nro_panels(df_equipos),
    }


def _build_finantial_pdf(
    cfg: dict,
    ws,
    dict_datos: dict,
    assets: dict,
    root_path: Path,
    output_pdf_path: Path,
    dynamic_layout: Optional[Dict[str, Any]] = None,
) -> Path:
    df_flujo_caja = cf_table(
        ws,
        column_cf_anio=cfg["parametros_busqueda_flujo_caja"]["column_anio"],
        column_cf_equipamiento=cfg["parametros_busqueda_flujo_caja"]["column_equipamiento"],
        column_cf_tarifa=cfg["parametros_busqueda_flujo_caja"]["column_tarifa"],
        column_cf_opex=cfg["parametros_busqueda_flujo_caja"]["column_OPEX"],
        column_cf_energia=cfg["parametros_busqueda_flujo_caja"]["column_energia"],
        column_cf_ahorro=cfg["parametros_busqueda_flujo_caja"]["column_ahorro"],
        column_cf_flujo_total=cfg["parametros_busqueda_flujo_caja"]["column_flujo_total"],
        column_cf_flujo_acumulado=cfg["parametros_busqueda_flujo_caja"]["column_flujo_acumulado"],
        maximo_vector=cfg["parametros_busqueda_flujo_caja"]["maximo_vector_flujo_caja"],
        minimo_vector=cfg["parametros_busqueda_flujo_caja"]["minimo_vector_flujo_caja"],
        dynamic_layout=dynamic_layout,
    )

    df_finantial_params = finantial_table(
        ws,
        column_cf_parametro=cfg["parametros_busqueda_parametros_financieros"]["column_parametro"],
        column_cf_valor=cfg["parametros_busqueda_parametros_financieros"]["column_valor"],
        column_cf_unidad=cfg["parametros_busqueda_parametros_financieros"]["column_unidad"],
        maximo_vector=cfg["parametros_busqueda_parametros_financieros"]["maximo_vector_parametros_financieros"],
        minimo_vector=cfg["parametros_busqueda_parametros_financieros"]["minimo_vector_parametros_financieros"],
        dynamic_layout=dynamic_layout,
    )
    print(df_finantial_params)

    output_files = generar_graficas(df_flujo_caja)

    build_reporte_pdf_finantial(
        output_pdf_path=output_pdf_path,
        root_path=root_path,
        assets_logo=assets["logo"],
        assets_firma=assets["firma"],
        DF_CF=df_flujo_caja,
        DF_FINANTIAL=df_finantial_params,
        DF_PLOTS=list(output_files.values()),
        dict_datos=dict_datos,
    )

    return output_pdf_path


def compute_finantial_metrics(
    cfg: dict,
    ws,
    dynamic_layout: Optional[Dict[str, Any]] = None,
) -> dict:
    df_flujo_caja = cf_table(
        ws,
        column_cf_anio=cfg["parametros_busqueda_flujo_caja"]["column_anio"],
        column_cf_equipamiento=cfg["parametros_busqueda_flujo_caja"]["column_equipamiento"],
        column_cf_tarifa=cfg["parametros_busqueda_flujo_caja"]["column_tarifa"],
        column_cf_opex=cfg["parametros_busqueda_flujo_caja"]["column_OPEX"],
        column_cf_energia=cfg["parametros_busqueda_flujo_caja"]["column_energia"],
        column_cf_ahorro=cfg["parametros_busqueda_flujo_caja"]["column_ahorro"],
        column_cf_flujo_total=cfg["parametros_busqueda_flujo_caja"]["column_flujo_total"],
        column_cf_flujo_acumulado=cfg["parametros_busqueda_flujo_caja"]["column_flujo_acumulado"],
        maximo_vector=cfg["parametros_busqueda_flujo_caja"]["maximo_vector_flujo_caja"],
        minimo_vector=cfg["parametros_busqueda_flujo_caja"]["minimo_vector_flujo_caja"],
        dynamic_layout=dynamic_layout,
    )

    if df_flujo_caja.empty:
        return {"LCOE": 0.0, "time_retorn": 0.0}

    lcoe = _safe_div(
        df_flujo_caja["Equipamiento"].sum() - df_flujo_caja["OPEX"].sum(),
        df_flujo_caja["Energía"].sum(),
    )
    valid_return_rows = df_flujo_caja[
        df_flujo_caja["Año"].notna() & df_flujo_caja["Flujo Acumulado"].notna()
    ]
    time_retorn = next(
        (
            int(anio) if float(anio).is_integer() else float(anio)
            for anio, flujo in zip(valid_return_rows["Año"], valid_return_rows["Flujo Acumulado"])
            if flujo >= 0
        ),
        0,
    )
    return {"LCOE": lcoe, "time_retorn": time_retorn}


def generate_report(
    report_type: ReportType,
    CFG: dict,
    ws,
    dict_datos: dict,
    assets: dict,
    output_pdf_path: Optional[Union[str, Path]] = None,
    dynamic_layout: Optional[Dict[str, Any]] = None,
) -> Path:
    if ws is None:
        if report_type == "quote":
            raise ValueError("Se requiere una hoja de Excel válida para generar la cotización.")
        raise ValueError("Se requiere una hoja de Excel válida para generar el análisis financiero.")

    root_path = _resolve_root_path(CFG)
    output_path = _resolve_output_pdf_path(root_path, output_pdf_path)

    if report_type == "quote":
        return _build_quote_pdf(
            CFG,
            ws,
            dict_datos,
            assets,
            root_path,
            output_path,
            dynamic_layout=dynamic_layout,
        )
    if report_type == "finantial":
        return _build_finantial_pdf(
            CFG,
            ws,
            dict_datos,
            assets,
            root_path,
            output_path,
            dynamic_layout=dynamic_layout,
        )

    raise ValueError(f"Tipo de reporte no soportado: {report_type}")



if __name__ == "__main__":
    # Útil para probar localmente el core sin API
    CFG = build_config()

    # ----------------------------------------------------
    # ROOT_PATH absoluto
    # ----------------------------------------------------
    root_cfg = CFG["data_roots"]["ROOT_PATH"]
    ROOT_PATH = Path(root_cfg).resolve()

    # ----------------------------------------------------
    # Cargar el excel principal del sistema
    # ----------------------------------------------------
    budget_cfg = CFG["data_roots"].get("Budget_File")
    if not budget_cfg:
        raise ValueError("CFG['data_roots']['Budget_File'] está vacío.")

    Budget_File = Path(budget_cfg)
    if not Budget_File.is_absolute():
        Budget_File = (ROOT_PATH / Budget_File).resolve()

    if not Budget_File.exists():
        raise FileNotFoundError(f"No se encontró el Budget: {Budget_File}")

    wb1 = load_workbook(Budget_File, data_only=True)
    ws_cotizacion = wb1["COTIZACIÓN"]
    ws_finanzas = wb1["FLUJO DE CAJA" if "FLUJO DE CAJA" in wb1.sheetnames else "ANÁLISIS FINANCIERO"]

    ### Declarar datos del reporte de cotizacion
    dict_datos = {
        "cliente": ws_cotizacion[f"{CFG['datos_informe']['column_cliente']}{CFG['datos_informe']['row_cliente']}"].value,
        "ruc_dni": ws_cotizacion[f"{CFG['datos_informe']['column_RUC']}{CFG['datos_informe']['row_RUC']}"].value,
        "proyecto": ws_cotizacion[f"{CFG['datos_informe']['column_Proyecto']}{CFG['datos_informe']['row_Proyecto']}"].value,
        "fecha": CFG["datos_informe"]["Fecha"],
        "lugar": ws_cotizacion[f"{CFG['datos_informe']['column_Lugar']}{CFG['datos_informe']['row_Lugar']}"].value,
        "atencion": ws_cotizacion[f"{CFG['datos_informe']['column_Atencion']}{CFG['datos_informe']['row_Atencion']}"].value,
    }

    assets = {
        "logo": ROOT_PATH / "src" / "assets" / "TEC_logo.png",
        "firma": ROOT_PATH / "src" / "assets" / "Firma_jguerrero.png",
    }

    pdf_path_quote = generate_report("quote", CFG, ws_cotizacion, dict_datos, assets)
    pdf_path_finantial = generate_report("finantial", CFG, ws_finanzas, dict_datos, assets)
    print(f"PDF generado en: {pdf_path_quote}")
    print(f"PDF generado en: {pdf_path_finantial}")

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence

from openpyxl.worksheet.worksheet import Worksheet


def normalize_text(value: Any) -> str:
    """Normaliza texto para comparaciones tolerantes."""
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text


def value_looks_numeric(value: Any) -> bool:
    """Determina si un valor parece ser numérico, incluso si es una cadena con formato."""
    if isinstance(value, (int, float)):
        return True
    if not isinstance(value, str):
        return False
    cleaned = re.sub(r"[^0-9,.\-]", "", value.strip())
    return bool(cleaned) and any(ch.isdigit() for ch in cleaned)

## --- clases y funciones para manejo de Excel ---
@dataclass(frozen=True)
class CellRef:
    row: int
    column: int
    coordinate: str
    value: Any


@dataclass(frozen=True)
class LabelMatch:
    alias: str
    cell: CellRef
    score: int


@dataclass(frozen=True)
class TableDetection:
    header_row: int
    columns: Dict[str, int]


def iter_non_empty_cells(ws: Worksheet) -> Iterator[CellRef]:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if normalize_text(cell.value) == "":
                continue
            yield CellRef(
                row=cell.row,
                column=cell.column,
                coordinate=cell.coordinate,
                value=cell.value,
            )


def find_label_cell(
    ws: Worksheet,
    aliases: Sequence[str],
    *,
    exact_first: bool = True,
) -> Optional[LabelMatch]:
    normalized_aliases = [normalize_text(alias) for alias in aliases if normalize_text(alias)]
    if not normalized_aliases:
        return None

    best_match: Optional[LabelMatch] = None
    for item in iter_non_empty_cells(ws):
        cell_text = normalize_text(item.value)
        if not cell_text:
            continue

        for alias in normalized_aliases:
            score = 0
            if exact_first and cell_text == alias:
                score = 100
            elif alias in cell_text:
                score = 70
            elif cell_text in alias:
                score = 60

            if score == 0:
                continue

            candidate = LabelMatch(alias=alias, cell=item, score=score)
            if best_match is None or candidate.score > best_match.score:
                best_match = candidate

    return best_match


def find_value_near_label(
    ws: Worksheet,
    label_match: LabelMatch,
    *,
    search_offsets: Optional[Sequence[tuple[int, int]]] = None,
    prefer_numeric: bool = False,
) -> Optional[CellRef]:
    offsets = list(
        search_offsets
        or [
            (0, 1),
            (0, 2),
            (1, 0),
            (1, 1),
            (-1, 0),
            (0, -1),
        ]
    )

    for row_offset, col_offset in offsets:
        row = label_match.cell.row + row_offset
        col = label_match.cell.column + col_offset
        if row < 1 or col < 1:
            continue

        value = ws.cell(row=row, column=col).value
        if value is None or normalize_text(value) == "":
            continue
        if prefer_numeric and not value_looks_numeric(value):
            continue

        return CellRef(
            row=row,
            column=col,
            coordinate=ws.cell(row=row, column=col).coordinate,
            value=value,
        )

    return None

## --- coordina búsqueda de etiquetas + búsqueda del valor cercano + fallback.
def resolve_named_values(
    ws: Worksheet,
    field_aliases: Dict[str, Sequence[str]],
    *,
    fallback: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    resolved: Dict[str, Any] = {}
    fallback = fallback or {}

    for field_name, aliases in field_aliases.items():
        label_match = find_label_cell(ws, aliases)
        if label_match is None:
            resolved[field_name] = fallback.get(field_name)
            continue

        value_cell = find_value_near_label(ws, label_match)
        if value_cell is None:
            resolved[field_name] = fallback.get(field_name)
            continue

        resolved[field_name] = value_cell.value

    return resolved


def row_matches_aliases(
    row_values: Sequence[Any],
    header_aliases: Dict[str, Sequence[str]],
) -> Optional[Dict[str, int]]:
    normalized_cells = [normalize_text(value) for value in row_values]
    columns: Dict[str, int] = {}

    for logical_name, aliases in header_aliases.items():
        normalized_aliases = [normalize_text(alias) for alias in aliases if normalize_text(alias)]
        found_index: Optional[int] = None

        for idx, cell_text in enumerate(normalized_cells, start=1):
            if not cell_text:
                continue
            if any(alias == cell_text or alias in cell_text for alias in normalized_aliases):
                found_index = idx
                break

        if found_index is None:
            return None

        columns[logical_name] = found_index

    return columns


def detect_table_headers(
    ws: Worksheet,
    header_aliases: Dict[str, Sequence[str]],
    *,
    min_hits: Optional[int] = None,
) -> Optional[TableDetection]:
    required_hits = min_hits or len(header_aliases)

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        match = row_matches_aliases(row, header_aliases)
        if match is not None and len(match) >= required_hits:
            return TableDetection(header_row=row_idx, columns=match)

    return None


def _row_has_meaningful_data(values: Iterable[Any]) -> bool:
    return any(value is not None and normalize_text(value) != "" for value in values)


def extract_table_rows(
    ws: Worksheet,
    detection: TableDetection,
    *,
    stop_after_blank_rows: int = 2,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    blank_streak = 0

    for row_idx in range(detection.header_row + 1, ws.max_row + 1):
        row_data = {
            logical_name: ws.cell(row=row_idx, column=column_idx).value
            for logical_name, column_idx in detection.columns.items()
        }

        if _row_has_meaningful_data(row_data.values()):
            rows.append(row_data)
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= stop_after_blank_rows:
                break

    return rows

## --- coordina detección de encabezados + extracción de filas + fallback.
def extract_table_with_fallback(
    ws: Worksheet,
    header_aliases: Dict[str, Sequence[str]],
    *,
    fallback_rows: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    detection = detect_table_headers(ws, header_aliases)
    if detection is None:
        return fallback_rows or []
    return extract_table_rows(ws, detection)

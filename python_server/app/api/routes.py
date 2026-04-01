# backend/app/api/routes.py
from __future__ import annotations

import json
import uuid
import tempfile
import shutil
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, Literal, Optional, Sequence

from fastapi import APIRouter, BackgroundTasks, Query, File
from fastapi import Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import select

from ..core.configs import build_config
from ..core.cotizaciones import generate_report, compute_finantial_metrics, compute_quote_metrics
from ..core.excel_local import detect_table_headers, resolve_named_values
from ..db.models import Project
from ..db.session import SessionLocal

router = APIRouter(tags=["reports"])
ReportType = Literal["quote", "finantial"]


QUOTE_FIELD_ALIASES: Dict[str, Sequence[str]] = {
    "cliente": ["cliente", "razon social", "empresa"],
    "ruc_dni": ["ruc", "dni", "ruc/dni", "ruc dni"],
    "proyecto": ["proyecto", "nombre del proyecto"],
    "lugar": ["lugar", "ubicacion", "direccion", "dirección"],
    "atencion": ["atencion", "atención", "contacto", "responsable"],
}

QUOTE_TABLE_ALIASES: Dict[str, Sequence[str]] = {
    "descripcion": ["descripcion", "descripción"],
    "unidad": ["unidad", "und"],
    "cantidad": ["cantidad", "cant"],
}

FINANCIAL_FLOW_TABLE_ALIASES: Dict[str, Sequence[str]] = {
    "anio": ["año", "anio"],
    "equipamiento": ["equipamiento", "inversion", "equipos"],
    "tarifa": ["tarifa", "Tarifa del cliente"],
    "opex": ["opex", "O&M"],
    "ahorro": ["ahorro", "ahorro en la facturación"],
    "flujo_total": ["flujo total"],
    "flujo_acumulado": ["flujo acumulado"],
}

FINANCIAL_PARAMS_TABLE_ALIASES: Dict[str, Sequence[str]] = {
    "parametro": ["parametro", "parámetro"],
    "valor": ["valor"],
    "unidad": ["unidad"],
}


def _normalize_sheet_name(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", name)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"[^a-zA-Z0-9]+", " ", normalized).strip().lower()
    return re.sub(r"\s+", " ", normalized)


def _find_sheet_by_aliases(sheetnames: Sequence[str], aliases: Sequence[str]) -> Optional[str]:
    normalized_aliases = [_normalize_sheet_name(alias) for alias in aliases]
    for sheet in sheetnames:
        normalized_sheet = _normalize_sheet_name(sheet)
        if normalized_sheet in normalized_aliases:
            return sheet
        if any(alias in normalized_sheet or normalized_sheet in alias for alias in normalized_aliases):
            return sheet
    return None


def _safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _project_root() -> Path:
    """
    Calcula el ROOT_PATH para tu core.
    Ajusta si mueves carpetas. Con esta estructura:
    backend/app/api/routes.py -> root = repo/
    """
    # routes.py -> api/ -> app/ -> backend/ -> repo/
    return Path(__file__).resolve().parents[3]


def _project_storage_dir(project_id: int) -> Path:
    storage_dir = _project_root() / "python_server" / "storage" / "projects" / str(project_id)
    storage_dir.mkdir(parents=True, exist_ok=True)
    return storage_dir


def _create_temp_workspace(background_tasks: BackgroundTasks) -> Path:
    temp_dir = tempfile.mkdtemp(prefix="cotizador_")
    temp_dir_path = Path(temp_dir)
    background_tasks.add_task(shutil.rmtree, temp_dir, ignore_errors=True)
    return temp_dir_path


def _build_report_context(cfg: Dict[str, Any]) -> Dict[str, Any]:
    root_path = Path(cfg["data_roots"]["ROOT_PATH"]).resolve()
    budget_file = Path(cfg["data_roots"]["Budget_File"]).resolve()

    try:
        workbook = load_workbook(budget_file, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo abrir el Excel cargado: {exc}") from exc

    sheet_aliases = {
        "cotizacion": [
            "COTIZACIÓN",
            "COTIZACION",
            "COTIZACIONES",
            "PRESUPUESTO",
            "PROPUESTA ECONOMICA",
        ],
        "finanzas": [
            "FLUJO DE CAJA",
            "FLUJO CAJA",
            "ANÁLISIS FINANCIERO",
            "ANALISIS FINANCIERO",
            "FINANZAS",
            "FINANCIERO",
            "CASH FLOW",
        ],
    }

    cotizacion_sheet_name = _find_sheet_by_aliases(workbook.sheetnames, sheet_aliases["cotizacion"])
    finanzas_sheet_name = _find_sheet_by_aliases(workbook.sheetnames, sheet_aliases["finanzas"])

    missing_sheets = []
    if not cotizacion_sheet_name:
        missing_sheets.append("COTIZACIÓN")
    if not finanzas_sheet_name:
        missing_sheets.append("FLUJO DE CAJA / ANÁLISIS FINANCIERO")

    if missing_sheets:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan hojas requeridas en el Excel: {', '.join(missing_sheets)}",
        )

    if cotizacion_sheet_name is None or finanzas_sheet_name is None:
        raise HTTPException(
            status_code=400,
            detail="Faltan hojas requeridas en el Excel",
        )

    ws_cotizacion = workbook[cotizacion_sheet_name]
    ws_finanzas = workbook[finanzas_sheet_name]
    datos_cfg = cfg["datos_informe"]

    fallback_dict_datos = {
        "cliente": ws_cotizacion[f"{datos_cfg['column_cliente']}{datos_cfg['row_cliente']}"].value,
        "ruc_dni": ws_cotizacion[f"{datos_cfg['column_RUC']}{datos_cfg['row_RUC']}"].value,
        "proyecto": ws_cotizacion[f"{datos_cfg['column_Proyecto']}{datos_cfg['row_Proyecto']}"].value,
        "lugar": ws_cotizacion[f"{datos_cfg['column_Lugar']}{datos_cfg['row_Lugar']}"].value,
        "atencion": ws_cotizacion[f"{datos_cfg['column_Atencion']}{datos_cfg['row_Atencion']}"].value,
    }
    dict_datos = resolve_named_values(
        ws_cotizacion,
        QUOTE_FIELD_ALIASES,
        fallback=fallback_dict_datos,
    )
    dict_datos["fecha"] = datos_cfg["Fecha"]

    dynamic_layout = {
        "quote_fields": dict_datos,
        "quote_table": detect_table_headers(ws_cotizacion, QUOTE_TABLE_ALIASES),
        "financial_flow_table": detect_table_headers(ws_finanzas, FINANCIAL_FLOW_TABLE_ALIASES),
        "financial_params_table": detect_table_headers(ws_finanzas, FINANCIAL_PARAMS_TABLE_ALIASES),
    }

    assets = {
        "logo": root_path / "client" / "src" / "assets" / "TEC_logo.png",
        "firma": root_path / "client" / "src" / "assets" / "Firma_jguerrero.png",
    }

    return {
        "ws_cotizacion": ws_cotizacion,
        "ws_finanzas": ws_finanzas,
        "dict_datos": dict_datos,
        "dynamic_layout": dynamic_layout,
        "assets": assets,
    }


def _detect_report_type(path: str) -> ReportType:
    if path.endswith("/reports/quote"):
        return "quote"
    if path.endswith("/reports/finantial"):
        return "finantial"
    raise HTTPException(status_code=404, detail="Ruta de reporte no soportada")


def _validate_excel_filename(filename: str) -> None:
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx/.xlsm/.xls")


def _parse_overrides(report_type: ReportType, overrides_json: Optional[str]) -> Dict[str, Any]:
    if report_type != "finantial" or not overrides_json:
        return {}

    try:
        overrides = json.loads(overrides_json)
        if not isinstance(overrides, dict):
            raise ValueError("overrides_json debe ser un JSON objeto (dict).")
        return overrides
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"overrides_json inválido: {exc}") from exc


def _save_uploaded_excel(project_id: int, filename: str) -> tuple[Path, Path]:
    project_dir = _project_storage_dir(project_id)
    excel_filename = Path(filename).name or "uploaded.xlsx"
    tmp_excel_path = project_dir / excel_filename
    return project_dir, tmp_excel_path


def _build_merged_overrides(root_path: Path, excel_path: Path, overrides: Dict[str, Any]) -> Dict[str, Any]:
    merged_overrides: Dict[str, Any] = {
        "data_roots": {
            "ROOT_PATH": str(root_path),
            "Budget_File": str(excel_path),
        }
    }

    if overrides:
        merged_overrides.update({k: v for k, v in overrides.items() if k != "data_roots"})
        if "data_roots" in overrides and isinstance(overrides["data_roots"], dict):
            merged_overrides["data_roots"].update(overrides["data_roots"])
            merged_overrides["data_roots"]["ROOT_PATH"] = str(root_path)
            merged_overrides["data_roots"]["Budget_File"] = str(excel_path)

    return merged_overrides


def _build_response_filename(report_type: ReportType) -> str:
    if report_type == "quote":
        return "reporte_final.pdf"
    return "reporte_finantial.pdf"

@router.get("/reports/download-excel")
async def download_excel(project_id: int = Query(...)):
    db = SessionLocal()
    try:
        project = db.execute(
            select(Project).where(Project.id == project_id)
        ).scalar_one_or_none()

        if project is None:
            raise HTTPException(status_code=404, detail="Project not found")

        if not project.excel_file_path:
            raise HTTPException(status_code=404, detail="El proyecto no tiene un Excel asociado")

        excel_path = Path(project.excel_file_path)

        if not excel_path.exists() or not excel_path.is_file():
            raise HTTPException(status_code=404, detail="El archivo Excel no existe en el servidor")

        return FileResponse(
            path=str(excel_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=excel_path.name,
        )
    finally:
        db.close()

@router.post("/reports/process-project")
async def process_project(
    project_id: int = Form(...),
    excel_file: UploadFile = File(...),
    overrides_json: Optional[str] = Form(None),
):
    filename = excel_file.filename or ""
    _validate_excel_filename(filename)
    root_path = _project_root()

    overrides: Dict[str, Any] = {}
    if overrides_json:
        try:
            overrides = json.loads(overrides_json)
            if not isinstance(overrides, dict):
                raise ValueError("overrides_json debe ser un JSON objeto (dict).")
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"overrides_json inválido: {exc}") from exc

    project_dir, tmp_excel_path = _save_uploaded_excel(project_id, filename)
    contents = await excel_file.read()
    tmp_excel_path.write_bytes(contents)

    merged_overrides = _build_merged_overrides(root_path, tmp_excel_path, overrides)
    cfg = build_config(merged_overrides)
    report_context = _build_report_context(cfg)

    request_id = uuid.uuid4().hex
    quote_pdf = project_dir / f"quote_{request_id}.pdf"
    finantial_pdf = project_dir / f"finantial_{request_id}.pdf"

    quote_pdf_path = generate_report(
        "quote",
        cfg,
        ws=report_context["ws_cotizacion"],
        dict_datos=report_context["dict_datos"],
        assets=report_context["assets"],
        output_pdf_path=quote_pdf,
        dynamic_layout=report_context["dynamic_layout"],
    )
    finantial_pdf_path = generate_report(
        "finantial",
        cfg,
        ws=report_context["ws_finanzas"],
        dict_datos=report_context["dict_datos"],
        assets=report_context["assets"],
        output_pdf_path=finantial_pdf,
        dynamic_layout=report_context["dynamic_layout"],
    )

    quote_metrics = compute_quote_metrics(
        cfg,
        report_context["ws_cotizacion"],
        dynamic_layout=report_context["dynamic_layout"],
    )
    finantial_metrics = compute_finantial_metrics(
        cfg,
        report_context["ws_finanzas"],
        dynamic_layout=report_context["dynamic_layout"],
    )
    if not quote_pdf_path.exists() or not finantial_pdf_path.exists():
        raise HTTPException(status_code=500, detail="No se generaron los PDFs esperados.")

    time_retorn = _safe_int(finantial_metrics.get("time_retorn"))

    db = SessionLocal()
    try:
        project = db.execute(select(Project).where(Project.id == project_id)).scalar_one_or_none()
        if project is None:
            raise HTTPException(status_code=404, detail="Project not found")
        project.project = str(report_context["dict_datos"]["cliente"])
        project.excel_file_path = str(tmp_excel_path)
        project.pdf_quote = str(quote_pdf_path)
        project.pdf_finantial = str(finantial_pdf_path)
        project.time_retorn = time_retorn
        project.nro_panels = int(quote_metrics["nro_panels"])
        project.price = float(quote_metrics["price"])
        db.commit()
    finally:
        db.close()

    return {
        "ok": True,
        "project_id": project_id,
        "excel_file": str(tmp_excel_path),
        "pdf_quote": str(quote_pdf_path),
        "pdf_finantial": str(finantial_pdf_path),
        "time_retorn": time_retorn,
        "nro_panels": int(quote_metrics["nro_panels"]),
        "price": float(quote_metrics["price"]),
    }


@router.post("/reports/quote")
@router.post("/reports/finantial")
async def create_report(
    request: Request,
    background_tasks: BackgroundTasks,
    excel_file: UploadFile = File(...),
    overrides_json: Optional[str] = Form(None),
):
    """
    Endpoint unificado para:
    - /reports/quote
    - /reports/finantial

    Recibe:
    - excel_file (xlsx) via multipart/form-data
    - overrides_json (opcional) via form field: string JSON (para finantial)

    Devuelve:
    - PDF generado (FileResponse)
    """

    report_type = _detect_report_type(request.url.path)

    filename = excel_file.filename or ""
    _validate_excel_filename(filename)

    root_path = _project_root()
    overrides = _parse_overrides(report_type, overrides_json)

    temp_dir_path = _create_temp_workspace(background_tasks)
    tmp_excel_path = temp_dir_path / Path(filename).name

    contents = await excel_file.read()
    tmp_excel_path.write_bytes(contents)

    merged_overrides = _build_merged_overrides(root_path, tmp_excel_path, overrides)
    cfg = build_config(merged_overrides)
    report_context = _build_report_context(cfg)

    request_id = uuid.uuid4().hex
    out_pdf = temp_dir_path / f"reporte_{request_id}.pdf"

    if report_type == "quote":
        pdf_path = generate_report(
            "quote",
            cfg,
            ws=report_context["ws_cotizacion"],
            dict_datos=report_context["dict_datos"],
            assets=report_context["assets"],
            output_pdf_path=out_pdf,
            dynamic_layout=report_context["dynamic_layout"],
        )
    else:
        pdf_path = generate_report(
            "finantial",
            cfg,
            ws=report_context["ws_finanzas"],
            dict_datos=report_context["dict_datos"],
            assets=report_context["assets"],
            output_pdf_path=out_pdf,
            dynamic_layout=report_context["dynamic_layout"],
        )

    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail=f"No se generó el PDF esperado: {pdf_path}")

    return FileResponse(
        path=str(pdf_path),
        media_type="application/pdf",
        filename=_build_response_filename(report_type),
        background=background_tasks,
    )
